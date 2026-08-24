#!/usr/bin/env python3
"""Create a human-reviewable WordPress-to-Astro image alt workbook.

The workbook is intentionally separate from the machine ledger. It puts one asset per row,
embeds a thumbnail, and shows the Astro path, source provenance, original alt, suggested/final
alt fields, code usage locations, and a manual decision column.
"""
from __future__ import annotations

import json
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageOps

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / 'public'
DOCS = ROOT / 'docs'
LEDGER = DOCS / 'wordpress-to-astro-image-alt-ledger.json'
ARTICLE_MEDIA = DOCS / 'wordpress-article-media-migration.json'
OUT = DOCS / 'Sunlit_Image_Alt_Manual_Review.xlsx'
THUMB_DIR = DOCS / '.alt-review-thumbnails'

HEAD_FILL = PatternFill('solid', fgColor='0F4D3A')
SUB_FILL = PatternFill('solid', fgColor='EAF3ED')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')
SUCCESS_FILL = PatternFill('solid', fgColor='E2F0D9')
WHITE_FONT = Font(color='FFFFFF', bold=True)
THIN = Side(style='thin', color='D9E2D7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def discover_usages(paths: list[str]) -> dict[str, list[str]]:
    usages: dict[str, list[str]] = defaultdict(list)
    sources = list((ROOT / 'src').rglob('*.astro')) + list((ROOT / 'src').rglob('*.ts')) + list((ROOT / 'src').rglob('*.md'))
    text_cache = {path: path.read_text(encoding='utf-8-sig', errors='ignore') for path in sources}
    for asset_path in paths:
        for source, text in text_cache.items():
            if asset_path in text:
                usages[asset_path].append(source.relative_to(ROOT).as_posix())
    return usages


def make_thumbnail(asset_path: Path, key: str) -> Path | None:
    if not asset_path.exists():
        return None
    try:
        with Image.open(asset_path) as raw:
            image = ImageOps.exif_transpose(raw)
            if image.mode in ('RGBA', 'LA'):
                background = Image.new('RGB', image.size, 'white')
                background.paste(image, mask=image.getchannel('A'))
                image = background
            else:
                image = image.convert('RGB')
            image.thumbnail((150, 104))
            output = THUMB_DIR / f'{key}.jpg'
            image.save(output, 'JPEG', quality=82, optimize=True)
            return output
    except Exception:
        return None


def style_sheet(sheet, widths: dict[str, float]) -> None:
    sheet.freeze_panes = 'A3'
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = HEAD_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    sheet.row_dimensions[1].height = 34
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.auto_filter.ref = sheet.dimensions


def add_image(sheet, row: int, thumbnail: Path | None) -> None:
    sheet.row_dimensions[row].height = 84
    if thumbnail is None:
        sheet.cell(row, 2, '缩略图不可用').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return
    picture = ExcelImage(str(thumbnail))
    picture.width = 112
    picture.height = 78
    sheet.add_image(picture, f'B{row}')


def add_row(sheet, row: int, values: list[object], thumbnail: Path | None, decision: str, fill: PatternFill | None) -> None:
    for column, value in enumerate(values, start=1):
        if column == 2:
            continue
        cell = sheet.cell(row, column, value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = BORDER
        if fill:
            cell.fill = fill
    sheet.cell(row, 2).border = BORDER
    if fill:
        sheet.cell(row, 2).fill = fill
    sheet.cell(row, 11, decision).alignment = Alignment(vertical='top', wrap_text=True)
    sheet.cell(row, 11).border = BORDER
    if fill:
        sheet.cell(row, 11).fill = fill
    add_image(sheet, row, thumbnail)


def main() -> None:
    if THUMB_DIR.exists():
        shutil.rmtree(THUMB_DIR)
    THUMB_DIR.mkdir(parents=True)
    ledger = json.loads(LEDGER.read_text(encoding='utf-8'))
    article_data = json.loads(ARTICLE_MEDIA.read_text(encoding='utf-8'))

    base_assets = ledger['assets']
    article_assets = []
    for article in article_data['articles']:
        if article.get('cover'):
            image = article['cover'].copy()
            image.update({'article_slug': article['slug'], 'role': '封面图'})
            article_assets.append(image)
        for index, image in enumerate(article.get('inline', []), start=1):
            copied = image.copy()
            copied.update({'article_slug': article['slug'], 'role': f'正文图 {index}'})
            article_assets.append(copied)

    all_paths = [asset['astro_public_path'] for asset in base_assets] + [asset['src'] for asset in article_assets]
    usages = discover_usages(all_paths)

    workbook = Workbook()
    overview = workbook.active
    overview.title = '使用说明'
    overview.sheet_view.showGridLines = False
    overview.merge_cells('A1:H1')
    overview['A1'] = 'Sunlit Solar Roof — WordPress → Astro 图片 Alt 人工复核工作簿'
    overview['A1'].fill = HEAD_FILL
    overview['A1'].font = Font(color='FFFFFF', bold=True, size=15)
    overview['A1'].alignment = Alignment(horizontal='center', vertical='center')
    overview.row_dimensions[1].height = 30
    overview.append([])
    notes = [
        ('本工作簿怎么用', '每一行只对应一张 Astro 图片。请先打开“待人工复核 (56)”工作表；此处的图片在 WordPress 导出中没有精确文件名匹配，不能声称继承了旧站 alt。'),
        ('缩略图', '第二列是实际保存在 Astro 项目中的图片缩略图，可直接目视确认内容。'),
        ('Astro 保存路径', '第三列是迁移后站点使用的 URL 路径；第四列是项目中的实际保存位置。'),
        ('WordPress 原 alt', '仅在文件名精确匹配且旧站存在有效 alt 时提供；它是迁移参考，不应因关键词不同而盲目保留。'),
        ('建议/最终 alt', '“当前建议 alt”显示旧站有效 alt 或当前安全兜底；请把你确认后的版本填写在“你的最终 alt”。'),
        ('处理决策', '请选择：保留 WordPress 原 alt、改用你的最终 alt、装饰性图片 alt=""、或待进一步确认。'),
        ('重要边界', '编辑 Excel 不会自动改网站代码。你完成审核后，把工作簿发回给我，我会将“你的最终 alt”和决策同步到 Astro 数据模块。'),
    ]
    overview.append(['字段/步骤', '说明'])
    for cell in overview[3]:
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
    for title, description in notes:
        overview.append([title, description])
    for row in overview.iter_rows(min_row=4, max_row=3 + len(notes), min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = BORDER
    overview.column_dimensions['A'].width = 25
    overview.column_dimensions['B'].width = 112
    overview.append([])
    overview.append(['覆盖摘要', '数量'])
    for cell in overview[12]:
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
    summary_rows = [
        ('Astro 原有本地图片', len(base_assets)),
        ('已精确继承有效 WordPress alt', sum(1 for item in base_assets if item['mapping_status'] == 'legacy_alt_inherited')),
        ('需人工复核：无精确 WordPress 文件匹配', sum(1 for item in base_assets if item['mapping_status'] != 'legacy_alt_inherited')),
        ('已恢复的文章图片', len(article_assets)),
        ('工作簿总图片行数', len(base_assets) + len(article_assets)),
    ]
    for title, count in summary_rows:
        overview.append([title, count])
    for row in overview.iter_rows(min_row=13, max_row=12 + len(summary_rows), min_col=1, max_col=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    headers = ['序号', '图片预览', 'Astro URL 路径', 'Astro 项目保存位置', '实际代码使用位置', 'WordPress 原上传路径', 'WordPress 原图片 alt', '迁移状态', '当前建议 alt', '你的最终 alt（请填）', '处理决策', '备注']
    decision_validation = DataValidation(type='list', formula1='"保留 WordPress 原 alt,改用你的最终 alt,装饰性图片 alt=\"\",待进一步确认"', allow_blank=False)

    review = workbook.create_sheet('待人工复核 (56)')
    review.append(headers)
    review_assets = [asset for asset in base_assets if asset['mapping_status'] != 'legacy_alt_inherited']
    for index, asset in enumerate(review_assets, start=1):
        suggested = ''
        values = [
            index, '', asset['astro_public_path'], asset['astro_file_path'], '\n'.join(usages.get(asset['astro_public_path'], ['未检测到直接引用'])),
            asset['wordpress_upload_path'] or '无精确 WordPress 匹配', asset['legacy_alt'] or '无',
            '需人工复核：无精确 WordPress 文件匹配', suggested, '', '待进一步确认', '',
        ]
        thumb = make_thumbnail(ROOT / asset['astro_file_path'], f'review-{index}')
        add_row(review, index + 1, values, thumb, '待进一步确认', WARN_FILL)
    style_sheet(review, {'A': 7, 'B': 20, 'C': 58, 'D': 64, 'E': 35, 'F': 40, 'G': 50, 'H': 29, 'I': 46, 'J': 46, 'K': 26, 'L': 30})
    review.add_data_validation(decision_validation)
    decision_validation.add(f'K2:K{len(review_assets)+1}')

    inherited = workbook.create_sheet('已继承 WordPress alt (138)')
    inherited.append(headers)
    inherited_assets = [asset for asset in base_assets if asset['mapping_status'] == 'legacy_alt_inherited']
    for index, asset in enumerate(inherited_assets, start=1):
        values = [
            index, '', asset['astro_public_path'], asset['astro_file_path'], '\n'.join(usages.get(asset['astro_public_path'], ['未检测到直接引用'])),
            asset['wordpress_upload_path'], asset['legacy_alt'], '已精确继承 WordPress alt', asset['legacy_alt'], asset['legacy_alt'], '保留 WordPress 原 alt', '',
        ]
        thumb = make_thumbnail(ROOT / asset['astro_file_path'], f'inherited-{index}')
        add_row(inherited, index + 1, values, thumb, '保留 WordPress 原 alt', SUCCESS_FILL)
    style_sheet(inherited, {'A': 7, 'B': 20, 'C': 58, 'D': 64, 'E': 35, 'F': 40, 'G': 55, 'H': 28, 'I': 55, 'J': 55, 'K': 26, 'L': 30})
    inherited.add_data_validation(decision_validation)
    decision_validation.add(f'K2:K{len(inherited_assets)+1}')

    articles = workbook.create_sheet(f'文章图片 ({len(article_assets)})')
    articles.append(headers)
    for index, asset in enumerate(article_assets, start=1):
        status = asset['altStatus']
        decision = '保留 WordPress 原 alt' if status == 'wordpress-inherited' else '待进一步确认'
        display_status = '已继承 WordPress alt' if status == 'wordpress-inherited' else '旧站 alt 为空/泛化：需人工复核'
        values = [
            index, '', asset['src'], f"public{asset['src']}", '\n'.join(usages.get(asset['src'], ['文章 Markdown 中引用'])),
            asset.get('wordpressSourceUrl', ''), asset.get('wordpressSourceAlt', ''), display_status,
            asset['alt'], asset['alt'], decision, f"文章：{asset['article_slug']}；角色：{asset['role']}；章节：{asset.get('afterHeading', '封面') or '封面'}",
        ]
        thumb = make_thumbnail(PUBLIC / asset['src'].lstrip('/'), f'article-{index}')
        add_row(articles, index + 1, values, thumb, decision, SUCCESS_FILL if status == 'wordpress-inherited' else WARN_FILL)
    style_sheet(articles, {'A': 7, 'B': 20, 'C': 58, 'D': 64, 'E': 35, 'F': 70, 'G': 55, 'H': 31, 'I': 55, 'J': 55, 'K': 26, 'L': 42})
    articles.add_data_validation(decision_validation)
    decision_validation.add(f'K2:K{len(article_assets)+1}')

    for sheet in (review, inherited, articles):
        sheet.conditional_formatting.add(f'H2:H{sheet.max_row}', FormulaRule(formula=[f'ISNUMBER(SEARCH("需人工",H2))'], fill=WARN_FILL))

    workbook.save(OUT)
    shutil.rmtree(THUMB_DIR)
    print(f'Wrote {OUT.relative_to(ROOT)}')
    print(f'Review: {len(review_assets)} | Inherited: {len(inherited_assets)} | Article: {len(article_assets)}')


if __name__ == '__main__':
    main()
